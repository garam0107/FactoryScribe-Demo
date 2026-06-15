import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def generate_quotation_xlsx(template_path: str, output_path: str, data: dict) -> str:
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb["견적서"]

    ws["B3"] = data["customer_name"]
    ws["B4"] = data["project_name"]
    ws["B5"] = data["quotation_date"]

    start_row = 9
    total = 0

    for idx, item in enumerate(data["items"]):
        row = start_row + idx

        ws[f"A{row}"] = item["item_name"]
        ws[f"B{row}"] = item.get("part_code", "")
        ws[f"C{row}"] = item["quantity"]
        ws[f"D{row}"] = item["unit_price"]
        ws[f"E{row}"] = item["amount"]

        ws[f"C{row}"].number_format = '#,##0'
        ws[f"D{row}"].number_format = '#,##0"원"'
        ws[f"E{row}"].number_format = '#,##0"원"'

        total += item["amount"]

    ws["E20"] = total
    ws["E20"].number_format = '#,##0"원"'

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return output_path