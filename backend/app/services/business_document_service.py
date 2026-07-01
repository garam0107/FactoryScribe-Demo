import os
import re
from datetime import date, datetime
from typing import Iterable

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.business_document import (
    PurchaseOrderDocument,
    PurchaseOrderItem,
    QuotationDocument,
    QuotationItem,
)
from app.models.repository import DocumentRepository
from app.utils.ids import new_id
from app.utils.time import now_utc


QUOTE_ITEM_START_ROW = 17
QUOTE_ITEM_END_ROW = 34
PURCHASE_ORDER_ITEM_START_ROW = 11


def _cell_text(value) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None

    return float(match.group(0))


def _parse_korean_date(text: str | None) -> date | None:
    if not text:
        return None

    match = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    if not match:
        return None

    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day)


def _to_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return _parse_korean_date(str(value))


def _extract_document_no(text: str | None, prefix: str) -> str | None:
    if not text:
        return None

    match = re.search(rf"({prefix}-\d{{4}}-\d+)", text)
    if match:
        return match.group(1)

    stripped = text.strip()
    return stripped if stripped.startswith(prefix) else None


def _extract_recipient_from_quote_cell(text: str | None) -> str | None:
    if not text:
        return None

    first_line = text.splitlines()[0].strip()
    if "귀하" in first_line:
        return first_line.split("귀하", 1)[0].strip()

    return first_line or None


def _extract_quote_item_name_and_code(text: str | None) -> tuple[str | None, str | None]:
    if not text:
        return None, None

    match = re.search(r"\(([^()]+)\)\s*$", text)
    if not match:
        return text.strip(), None

    item_code = match.group(1).strip()
    item_name = text[: match.start()].strip()
    return item_name or None, item_code or None


def _extract_purchase_order_item_parts(text: str | None) -> tuple[str | None, str | None, str | None]:
    if not text:
        return None, None, None

    parts = [part.strip() for part in text.split("/") if part.strip()]
    item_name = parts[0] if len(parts) >= 1 else None
    spec = parts[1] if len(parts) >= 2 else None
    item_code = parts[2] if len(parts) >= 3 else None

    return item_name, spec, item_code


def _extract_unit_price_from_note(text: str | None) -> float | None:
    if not text:
        return None

    match = re.search(r"단가\s*([\d,]+(?:\.\d+)?)\s*원", text)
    if not match:
        return None

    return _to_float(match.group(1))


def _extract_status_from_note(text: str | None) -> str | None:
    if not text or "/" not in text:
        return None

    status = text.rsplit("/", 1)[-1].strip()
    return status or None


def _iter_xlsx_files(root_path: str) -> Iterable[str]:
    for root, _, filenames in os.walk(root_path):
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            if not filename.lower().endswith(".xlsx"):
                continue
            yield os.path.join(root, filename)


def _is_quotation_sheet(file_path: str, ws) -> bool:
    filename = os.path.basename(file_path)
    quotation_no = _extract_document_no(_cell_text(ws["B2"].value), "QT")
    return bool(quotation_no or "견적서" in filename or _cell_text(ws["B3"].value) == "견적서")


def _is_purchase_order_sheet(file_path: str, ws) -> bool:
    filename = os.path.basename(file_path)
    purchase_order_no = _extract_document_no(_cell_text(ws["B2"].value), "PO")
    return bool(purchase_order_no or "발주서" in filename or _cell_text(ws["H1"].value) == "발 주 서")


def _parse_quotation_sheet(
    repository_id: str,
    file_path: str,
    ws,
    now: datetime,
) -> tuple[QuotationDocument, list[QuotationItem]] | None:
    quotation_no = _extract_document_no(_cell_text(ws["B2"].value), "QT")
    if not quotation_no:
        return None

    document_id = new_id("quote_doc")
    document = QuotationDocument(
        id=document_id,
        repository_id=repository_id,
        quotation_no=quotation_no,
        quotation_date=_parse_korean_date(_cell_text(ws["B5"].value)),
        recipient_company_name=_extract_recipient_from_quote_cell(_cell_text(ws["B5"].value)),
        issuer_company_name=_cell_text(ws["I3"].value),
        project_name=_cell_text(ws["C8"].value),
        delivery_terms=_cell_text(ws["C9"].value),
        payment_terms=_cell_text(ws["C10"].value),
        valid_until_text=_cell_text(ws["C11"].value),
        source_filename=os.path.basename(file_path),
        source_sheet_name=ws.title,
        created_at=now,
        updated_at=now,
    )

    items: list[QuotationItem] = []
    for row_idx in range(QUOTE_ITEM_START_ROW, QUOTE_ITEM_END_ROW + 1):
        raw_item = _cell_text(ws[f"B{row_idx}"].value)
        quantity = _to_float(ws[f"H{row_idx}"].value)
        unit_price = _to_float(ws[f"J{row_idx}"].value)

        item_name, item_code = _extract_quote_item_name_and_code(raw_item)
        if not item_name or not item_code or quantity is None:
            continue

        supply_amount = quantity * unit_price if unit_price is not None else None
        tax_amount = supply_amount * 0.1 if supply_amount is not None else None
        total_amount = supply_amount + tax_amount if supply_amount is not None and tax_amount is not None else None

        items.append(
            QuotationItem(
                id=new_id("quote_item"),
                quotation_document_id=document_id,
                repository_id=repository_id,
                item_code=item_code,
                item_name=item_name,
                spec=_cell_text(ws[f"E{row_idx}"].value),
                quantity=quantity,
                unit_price=unit_price,
                supply_amount=supply_amount,
                tax_amount=tax_amount,
                total_amount=total_amount,
                source_row=row_idx,
                created_at=now,
                updated_at=now,
            )
        )

    if not items:
        return None

    return document, items


def _parse_purchase_order_sheet(
    repository_id: str,
    file_path: str,
    ws,
    now: datetime,
) -> tuple[PurchaseOrderDocument, list[PurchaseOrderItem]] | None:
    purchase_order_no = _extract_document_no(_cell_text(ws["B2"].value), "PO")
    if not purchase_order_no:
        return None

    document_id = new_id("po_doc")
    document = PurchaseOrderDocument(
        id=document_id,
        repository_id=repository_id,
        purchase_order_no=purchase_order_no,
        order_date=_parse_korean_date(_cell_text(ws["O3"].value)),
        recipient_company_name=_cell_text(ws["O5"].value),
        issuer_company_name=_cell_text(ws["D5"].value),
        project_name=_cell_text(ws["I7"].value),
        issuer_contact_name=_cell_text(ws["J5"].value),
        issuer_contact_text=_cell_text(ws["D8"].value),
        source_filename=os.path.basename(file_path),
        source_sheet_name=ws.title,
        created_at=now,
        updated_at=now,
    )

    items: list[PurchaseOrderItem] = []
    for row_idx in range(PURCHASE_ORDER_ITEM_START_ROW, ws.max_row + 1):
        row_no = _to_float(ws[f"A{row_idx}"].value)
        raw_item = _cell_text(ws[f"C{row_idx}"].value)
        quantity = _to_float(ws[f"K{row_idx}"].value)

        if row_no is None or not raw_item or quantity is None:
            continue

        item_name, spec, item_code = _extract_purchase_order_item_parts(raw_item)
        if not item_name or not item_code:
            continue

        note = _cell_text(ws[f"R{row_idx}"].value)
        items.append(
            PurchaseOrderItem(
                id=new_id("po_item"),
                purchase_order_document_id=document_id,
                repository_id=repository_id,
                item_code=item_code,
                item_name=item_name,
                spec=spec,
                unit=_cell_text(ws[f"H{row_idx}"].value),
                quantity=quantity,
                requested_delivery_date=_to_date(ws[f"N{row_idx}"].value),
                unit_price=_extract_unit_price_from_note(note),
                status_text=_extract_status_from_note(note),
                note=note,
                source_row=row_idx,
                created_at=now,
                updated_at=now,
            )
        )

    if not items:
        return None

    return document, items


def _replace_repository_business_documents(
    session: Session,
    repository_id: str,
    quotation_documents: list[QuotationDocument],
    quotation_items: list[QuotationItem],
    purchase_order_documents: list[PurchaseOrderDocument],
    purchase_order_items: list[PurchaseOrderItem],
) -> None:
    for row in session.exec(
        select(QuotationItem).where(QuotationItem.repository_id == repository_id)
    ).all():
        session.delete(row)

    for row in session.exec(
        select(QuotationDocument).where(QuotationDocument.repository_id == repository_id)
    ).all():
        session.delete(row)

    for row in session.exec(
        select(PurchaseOrderItem).where(PurchaseOrderItem.repository_id == repository_id)
    ).all():
        session.delete(row)

    for row in session.exec(
        select(PurchaseOrderDocument).where(PurchaseOrderDocument.repository_id == repository_id)
    ).all():
        session.delete(row)

    session.add_all(quotation_documents)
    session.add_all(quotation_items)
    session.add_all(purchase_order_documents)
    session.add_all(purchase_order_items)
    session.commit()


def sync_business_documents(session: Session, repository_id: str) -> dict:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    now = now_utc()
    quotation_documents: list[QuotationDocument] = []
    quotation_items: list[QuotationItem] = []
    purchase_order_documents: list[PurchaseOrderDocument] = []
    purchase_order_items: list[PurchaseOrderItem] = []
    matched_files: set[str] = set()

    for file_path in _iter_xlsx_files(repo.path):
        workbook = load_workbook(file_path, data_only=True)

        for ws in workbook.worksheets:
            if _is_quotation_sheet(file_path, ws):
                parsed_quotation = _parse_quotation_sheet(repository_id, file_path, ws, now)
                if parsed_quotation:
                    document, items = parsed_quotation
                    quotation_documents.append(document)
                    quotation_items.extend(items)
                    matched_files.add(os.path.basename(file_path))
                continue

            if _is_purchase_order_sheet(file_path, ws):
                parsed_purchase_order = _parse_purchase_order_sheet(repository_id, file_path, ws, now)
                if parsed_purchase_order:
                    document, items = parsed_purchase_order
                    purchase_order_documents.append(document)
                    purchase_order_items.extend(items)
                    matched_files.add(os.path.basename(file_path))

    if not quotation_documents and not purchase_order_documents:
        raise ValueError("quotation or purchase order worksheet not found in repository")

    _replace_repository_business_documents(
        session=session,
        repository_id=repository_id,
        quotation_documents=quotation_documents,
        quotation_items=quotation_items,
        purchase_order_documents=purchase_order_documents,
        purchase_order_items=purchase_order_items,
    )

    return {
        "repository_id": repository_id,
        "files": sorted(matched_files),
        "quotation_documents": len(quotation_documents),
        "quotation_items": len(quotation_items),
        "purchase_order_documents": len(purchase_order_documents),
        "purchase_order_items": len(purchase_order_items),
    }


def _quotation_document_to_dict(
    document: QuotationDocument,
    items: list[QuotationItem],
) -> dict:
    data = _model_to_dict(document)
    data["items"] = [_model_to_dict(item) for item in items]
    return data


def _purchase_order_document_to_dict(
    document: PurchaseOrderDocument,
    items: list[PurchaseOrderItem],
) -> dict:
    data = _model_to_dict(document)
    data["items"] = [_model_to_dict(item) for item in items]
    return data


def _model_to_dict(model) -> dict:
    if hasattr(model, "model_dump"):
        return model.model_dump()

    return model.dict()


def list_quotation_documents(session: Session, repository_id: str) -> list[dict]:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    documents = session.exec(
        select(QuotationDocument)
        .where(QuotationDocument.repository_id == repository_id)
        .order_by(QuotationDocument.quotation_date.desc(), QuotationDocument.quotation_no.desc())
    ).all()

    if not documents:
        raise ValueError("quotation documents not loaded")

    results = []
    for document in documents:
        items = session.exec(
            select(QuotationItem)
            .where(QuotationItem.quotation_document_id == document.id)
            .order_by(QuotationItem.source_row.asc())
        ).all()
        results.append(_quotation_document_to_dict(document, items))

    return results


def get_quotation_document(
    session: Session,
    repository_id: str,
    quotation_document_id: str,
) -> dict:
    document = session.get(QuotationDocument, quotation_document_id)
    if not document or document.repository_id != repository_id:
        raise ValueError("quotation document not found")

    items = session.exec(
        select(QuotationItem)
        .where(QuotationItem.quotation_document_id == document.id)
        .order_by(QuotationItem.source_row.asc())
    ).all()

    return _quotation_document_to_dict(document, items)


def list_purchase_order_documents(session: Session, repository_id: str) -> list[dict]:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    documents = session.exec(
        select(PurchaseOrderDocument)
        .where(PurchaseOrderDocument.repository_id == repository_id)
        .order_by(PurchaseOrderDocument.order_date.desc(), PurchaseOrderDocument.purchase_order_no.desc())
    ).all()

    if not documents:
        raise ValueError("purchase order documents not loaded")

    results = []
    for document in documents:
        items = session.exec(
            select(PurchaseOrderItem)
            .where(PurchaseOrderItem.purchase_order_document_id == document.id)
            .order_by(PurchaseOrderItem.source_row.asc())
        ).all()
        results.append(_purchase_order_document_to_dict(document, items))

    return results


def get_purchase_order_document(
    session: Session,
    repository_id: str,
    purchase_order_document_id: str,
) -> dict:
    document = session.get(PurchaseOrderDocument, purchase_order_document_id)
    if not document or document.repository_id != repository_id:
        raise ValueError("purchase order document not found")

    items = session.exec(
        select(PurchaseOrderItem)
        .where(PurchaseOrderItem.purchase_order_document_id == document.id)
        .order_by(PurchaseOrderItem.source_row.asc())
    ).all()

    return _purchase_order_document_to_dict(document, items)
