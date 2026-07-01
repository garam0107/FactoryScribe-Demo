import os
from datetime import datetime

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.inventory_item import InventoryItem
from app.models.repository import DocumentRepository
from app.utils.ids import new_id
from app.utils.time import now_utc


HEADER_ALIASES = {
    "item_code": ["품목코드", "item code", "part number", "part no"],
    "item_name": ["품목명", "item name", "part name", "material name"],
    "category": ["카테고리", "category"],
    "spec": ["규격", "spec", "specification"],
    "unit": ["단위", "unit", "uom"],
    "supplier": ["공급사", "supplier", "vendor"],
    "current_stock": ["현재고", "현재 수량", "current stock", "stock", "on hand"],
    "safety_stock": ["안전재고", "safety stock", "minimum stock"],
    "target_stock": ["적정재고", "target stock", "optimal stock"],
    "avg_monthly_usage": ["평균월사용량", "average monthly usage", "avg monthly usage"],
    "current_unit_price": ["현재단가", "current unit price", "current price", "unit price"],
    "previous_unit_price": ["전월단가", "previous unit price", "last month price"],
    "price_change_rate": ["가격상승률", "price change rate", "price increase rate"],
    "stock_status": ["재고상태", "stock status"],
    "expected_depletion_days": ["예상소진일", "expected depletion days", "days to depletion"],
    "warehouse_location": ["창고위치", "warehouse location", "location"],
    "last_inbound_date": ["최종입고일", "last inbound date", "last receipt date"],
    "note": ["비고", "note", "remarks"],
}

REQUIRED_HEADERS = {
    "item_code",
    "item_name",
    "current_stock",
    "target_stock",
    "current_unit_price",
    "previous_unit_price",
    "price_change_rate",
}


def _normalize_header(value) -> str:
    if value is None:
        return ""

    return " ".join(str(value).strip().lower().replace("_", " ").split())


def _find_header_row(rows: list[tuple]) -> int | None:
    for idx, row in enumerate(rows[:20]):
        normalized = {_normalize_header(v) for v in row if _normalize_header(v)}
        hit_count = 0

        for aliases in HEADER_ALIASES.values():
            if any(alias in normalized for alias in aliases):
                hit_count += 1

        if hit_count >= 4:
            return idx

    return None


def _build_header_map(header_row: tuple) -> dict[str, int]:
    header_map: dict[str, int] = {}

    for idx, value in enumerate(header_row):
        normalized = _normalize_header(value)
        if not normalized:
            continue

        for key, aliases in HEADER_ALIASES.items():
            if key in header_map:
                continue
            if normalized in aliases:
                header_map[key] = idx
                break

    return header_map


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    return float(text)


def _to_datetime(value) -> datetime | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _iter_inventory_files(root_path: str) -> list[str]:
    files = []

    for root, _, filenames in os.walk(root_path):
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            if not filename.lower().endswith(".xlsx"):
                continue
            files.append(os.path.join(root, filename))

    return files


def _replace_repository_inventory_items(
    session: Session,
    repository_id: str,
    items: list[InventoryItem],
) -> None:
    existing_rows = session.exec(
        select(InventoryItem).where(InventoryItem.repository_id == repository_id)
    ).all()

    for row in existing_rows:
        session.delete(row)

    session.add_all(items)
    session.commit()


def _is_shortage_item(item: InventoryItem) -> bool:
    return item.target_stock is not None and item.current_stock < item.target_stock


def sync_inventory_items(session: Session, repository_id: str) -> dict:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    now = now_utc()
    imported_items: list[InventoryItem] = []
    matched_files: list[str] = []
    matched_sheets: list[str] = []

    for file_path in _iter_inventory_files(repo.path):
        workbook = load_workbook(file_path, data_only=True)
        matched_file = False

        for ws in workbook.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = _find_header_row(rows)
            if header_idx is None:
                continue

            header_map = _build_header_map(rows[header_idx])
            if not REQUIRED_HEADERS.issubset(header_map.keys()):
                continue

            matched_file = True
            matched_sheets.append(f"{os.path.basename(file_path)} / {ws.title}")

            for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                item_code = row[header_map["item_code"]] if header_map.get("item_code") is not None else None
                item_name = row[header_map["item_name"]] if header_map.get("item_name") is not None else None

                if item_code is None or item_name is None:
                    continue

                item_code_text = str(item_code).strip()
                item_name_text = str(item_name).strip()
                if not item_code_text or not item_name_text:
                    continue

                imported_items.append(
                    InventoryItem(
                        id=new_id("inv"),
                        repository_id=repository_id,
                        item_code=item_code_text,
                        item_name=item_name_text,
                        category=str(row[header_map["category"]]).strip() if "category" in header_map and row[header_map["category"]] is not None else None,
                        spec=str(row[header_map["spec"]]).strip() if "spec" in header_map and row[header_map["spec"]] is not None else None,
                        unit=str(row[header_map["unit"]]).strip() if "unit" in header_map and row[header_map["unit"]] is not None else None,
                        supplier=str(row[header_map["supplier"]]).strip() if "supplier" in header_map and row[header_map["supplier"]] is not None else None,
                        current_stock=_to_float(row[header_map["current_stock"]]) or 0,
                        safety_stock=_to_float(row[header_map["safety_stock"]]) if "safety_stock" in header_map else None,
                        target_stock=_to_float(row[header_map["target_stock"]]) if "target_stock" in header_map else None,
                        avg_monthly_usage=_to_float(row[header_map["avg_monthly_usage"]]) if "avg_monthly_usage" in header_map else None,
                        current_unit_price=_to_float(row[header_map["current_unit_price"]]) if "current_unit_price" in header_map else None,
                        previous_unit_price=_to_float(row[header_map["previous_unit_price"]]) if "previous_unit_price" in header_map else None,
                        price_change_rate=_to_float(row[header_map["price_change_rate"]]) if "price_change_rate" in header_map else None,
                        stock_status=str(row[header_map["stock_status"]]).strip() if "stock_status" in header_map and row[header_map["stock_status"]] is not None else None,
                        expected_depletion_days=_to_float(row[header_map["expected_depletion_days"]]) if "expected_depletion_days" in header_map else None,
                        warehouse_location=str(row[header_map["warehouse_location"]]).strip() if "warehouse_location" in header_map and row[header_map["warehouse_location"]] is not None else None,
                        last_inbound_date=_to_datetime(row[header_map["last_inbound_date"]]) if "last_inbound_date" in header_map else None,
                        note=str(row[header_map["note"]]).strip() if "note" in header_map and row[header_map["note"]] is not None else None,
                        source_filename=os.path.basename(file_path),
                        source_sheet_name=ws.title,
                        source_row=row_idx,
                        created_at=now,
                        updated_at=now,
                    )
                )

        if matched_file:
            matched_files.append(os.path.basename(file_path))

    if not imported_items:
        raise ValueError("inventory worksheet not found in repository")

    _replace_repository_inventory_items(session, repository_id, imported_items)

    return {
        "repository_id": repository_id,
        "files": matched_files,
        "sheets": matched_sheets,
        "imported_items": len(imported_items),
    }


def list_inventory_items(
    session: Session,
    repository_id: str,
    shortage_only: bool = False,
) -> list[dict]:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    items = session.exec(
        select(InventoryItem)
        .where(InventoryItem.repository_id == repository_id)
        .order_by(InventoryItem.created_at.desc(), InventoryItem.item_name.asc())
    ).all()

    if not items:
        raise ValueError("inventory items not loaded")

    results = []
    for item in items:
        is_shortage = _is_shortage_item(item)
        if shortage_only and not is_shortage:
            continue

        results.append(
            {
                "id": item.id,
                "repository_id": item.repository_id,
                "item_code": item.item_code,
                "item_name": item.item_name,
                "category": item.category,
                "spec": item.spec,
                "unit": item.unit,
                "supplier": item.supplier,
                "current_stock": item.current_stock,
                "safety_stock": item.safety_stock,
                "target_stock": item.target_stock,
                "avg_monthly_usage": item.avg_monthly_usage,
                "current_unit_price": item.current_unit_price,
                "previous_unit_price": item.previous_unit_price,
                "price_change_rate": item.price_change_rate,
                "stock_status": item.stock_status,
                "expected_depletion_days": item.expected_depletion_days,
                "warehouse_location": item.warehouse_location,
                "last_inbound_date": item.last_inbound_date,
                "note": item.note,
                "source_filename": item.source_filename,
                "source_sheet_name": item.source_sheet_name,
                "source_row": item.source_row,
                "created_at": item.created_at,
                "updated_at": item.updated_at,
                "is_shortage": is_shortage,
            }
        )

    return results


def get_inventory_dashboard(session: Session, repository_id: str) -> dict:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    items = session.exec(
        select(InventoryItem).where(InventoryItem.repository_id == repository_id)
    ).all()

    if not items:
        raise ValueError("inventory items not loaded")

    total_current_stock = sum(item.current_stock for item in items)
    total_target_stock = sum(item.target_stock or 0 for item in items)

    price_change_rates = [
        item.price_change_rate
        for item in items
        if item.price_change_rate is not None
    ]

    shortage_items = sum(
        1
        for item in items
        if _is_shortage_item(item)
    )

    inventory_remaining_rate = None
    if total_target_stock > 0:
        inventory_remaining_rate = (total_current_stock / total_target_stock) * 100

    average_price_increase_rate = None
    if price_change_rates:
        average_price_increase_rate = (
            sum(price_change_rates) / len(price_change_rates)
        ) * 100

    return {
        "repository_id": repository_id,
        "total_items": len(items),
        "total_current_stock": round(total_current_stock, 3),
        "total_target_stock": round(total_target_stock, 3),
        "inventory_remaining_rate": round(inventory_remaining_rate, 2) if inventory_remaining_rate is not None else None,
        "average_price_increase_rate": round(average_price_increase_rate, 2) if average_price_increase_rate is not None else None,
        "shortage_items": shortage_items,
    }
