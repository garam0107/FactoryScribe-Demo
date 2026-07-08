import os
from datetime import datetime

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.repository import DocumentRepository
from app.models.transaction_record import TransactionRecord
from app.utils.ids import new_id
from app.utils.time import now_utc


HEADER_ALIASES = {
    "transaction_date": ["거래일자", "transaction date", "date"],
    "transaction_no": ["거래번호", "transaction no", "transaction number"],
    "item_code": ["품목코드", "item code", "part number", "part no"],
    "item_name": ["품목명", "item name", "part name", "material name"],
    "transaction_type": ["거래유형", "transaction type", "type"],
    "quantity": ["수량", "quantity", "qty"],
    "unit_price": ["단가", "unit price", "price"],
    "amount": ["금액", "amount", "total amount"],
    "partner_name": ["거래처/부서", "거래처", "부서", "partner", "vendor", "customer"],
    "project_or_order_no": [
        "프로젝트/발주번호",
        "프로젝트",
        "발주번호",
        "project/order no",
        "project",
        "order no",
    ],
    "warehouse_location": ["창고위치", "창고", "warehouse location", "location"],
    "manager_name": ["담당자", "manager", "owner"],
}

# The demo transaction sheet is column-oriented, so a header map is safer than
# fixed cell addresses and still works when a few rows are added above the table.
REQUIRED_HEADERS = {
    "transaction_date",
    "item_name",
    "transaction_type",
    "quantity",
    "partner_name",
}


def _normalize_header(value) -> str:
    if value is None:
        return ""

    return " ".join(str(value).strip().lower().replace("_", " ").split())


def _cell_text(value) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _to_datetime(value) -> datetime | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    return None


def _iter_xlsx_files(root_path: str):
    for root, _, filenames in os.walk(root_path):
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            if not filename.lower().endswith(".xlsx"):
                continue
            yield os.path.join(root, filename)


def _find_header_row(rows: list[tuple]) -> int | None:
    for idx, row in enumerate(rows[:20]):
        normalized = {_normalize_header(v) for v in row if _normalize_header(v)}
        hit_count = 0

        for aliases in HEADER_ALIASES.values():
            if any(alias in normalized for alias in aliases):
                hit_count += 1

        if hit_count >= 5:
            return idx

    return None


def _build_header_map(header_row: tuple) -> dict[str, int]:
    header_map: dict[str, int] = {}

    for idx, value in enumerate(header_row):
        normalized = _normalize_header(value)
        if not normalized:
            continue

        for key, aliases in HEADER_ALIASES.items():
            if key in header_map:
                continue
            if normalized in aliases:
                header_map[key] = idx
                break

    return header_map


def _row_value(row: tuple, header_map: dict[str, int], key: str):
    index = header_map.get(key)
    if index is None or index >= len(row):
        return None
    return row[index]


def _replace_repository_transaction_records(
    session: Session,
    repository_id: str,
    records: list[TransactionRecord],
) -> None:
    # Sync is repository-scoped replacement, matching inventory/business document sync.
    existing_rows = session.exec(
        select(TransactionRecord).where(TransactionRecord.repository_id == repository_id)
    ).all()

    for row in existing_rows:
        session.delete(row)

    session.add_all(records)
    session.commit()


def sync_transaction_records(session: Session, repository_id: str) -> dict:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    now = now_utc()
    imported_records: list[TransactionRecord] = []
    matched_files: list[str] = []
    matched_sheets: list[str] = []

    for file_path in _iter_xlsx_files(repo.path):
        workbook = load_workbook(file_path, data_only=True)
        matched_file = False

        for ws in workbook.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = _find_header_row(rows)
            if header_idx is None:
                continue

            header_map = _build_header_map(rows[header_idx])
            if not REQUIRED_HEADERS.issubset(header_map.keys()):
                continue

            # Only worksheets that match the transaction schema are imported.
            matched_file = True
            matched_sheets.append(f"{os.path.basename(file_path)} / {ws.title}")

            for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                item_name = _cell_text(_row_value(row, header_map, "item_name"))
                partner_name = _cell_text(_row_value(row, header_map, "partner_name"))
                quantity = _to_float(_row_value(row, header_map, "quantity"))

                if not item_name or not partner_name or quantity is None:
                    continue

                imported_records.append(
                    TransactionRecord(
                        id=new_id("trx"),
                        repository_id=repository_id,
                        transaction_date=_to_datetime(
                            _row_value(row, header_map, "transaction_date")
                        ),
                        transaction_no=_cell_text(
                            _row_value(row, header_map, "transaction_no")
                        ),
                        item_code=_cell_text(_row_value(row, header_map, "item_code")),
                        item_name=item_name,
                        transaction_type=_cell_text(
                            _row_value(row, header_map, "transaction_type")
                        ),
                        quantity=quantity,
                        unit_price=_to_float(_row_value(row, header_map, "unit_price")),
                        amount=_to_float(_row_value(row, header_map, "amount")),
                        partner_name=partner_name,
                        project_or_order_no=_cell_text(
                            _row_value(row, header_map, "project_or_order_no")
                        ),
                        warehouse_location=_cell_text(
                            _row_value(row, header_map, "warehouse_location")
                        ),
                        manager_name=_cell_text(
                            _row_value(row, header_map, "manager_name")
                        ),
                        source_filename=os.path.basename(file_path),
                        source_sheet_name=ws.title,
                        source_row=row_idx,
                        created_at=now,
                        updated_at=now,
                    )
                )

        if matched_file:
            matched_files.append(os.path.basename(file_path))

    if not imported_records:
        raise ValueError("transaction worksheet not found in repository")

    _replace_repository_transaction_records(session, repository_id, imported_records)

    return {
        "repository_id": repository_id,
        "files": matched_files,
        "sheets": matched_sheets,
        "imported_records": len(imported_records),
    }


def list_transaction_records(session: Session, repository_id: str) -> list[TransactionRecord]:
    repo = session.get(DocumentRepository, repository_id)
    if not repo:
        raise ValueError("repository not found")

    return session.exec(
        select(TransactionRecord)
        .where(TransactionRecord.repository_id == repository_id)
        .order_by(TransactionRecord.transaction_date.desc(), TransactionRecord.source_row.asc())
    ).all()
