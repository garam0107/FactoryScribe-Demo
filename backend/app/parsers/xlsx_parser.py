from openpyxl import load_workbook
import os
from app.config import settings


def _cell_to_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _is_empty_row(row: tuple) -> bool:
    return not any(_cell_to_text(v) for v in row)


def _detect_header_row(rows: list[tuple]) -> int:
    header_keywords = {
        "제품명", "품명", "품번", "제품코드", "부품코드",
        "현재고", "재고", "수량", "안전재고", "창고",
        "단가", "거래처", "공급사", "자재명", "규격",
        "구분", "항목", "단위", "금액"
    }

    for idx, row in enumerate(rows):
        values = [_cell_to_text(v) for v in row]
        hit_count = sum(1 for v in values if v in header_keywords)

        if hit_count >= 2:
            return idx

    for idx, row in enumerate(rows):
        if not _is_empty_row(row):
            return idx

    return 0


def _detect_sheet_type(filename: str, sheet_name: str, headers: list[str]) -> str:
    lower_text = " ".join([filename, sheet_name, *headers]).lower()
    header_set = set(headers)

    inventory_keywords = {"현재고", "재고", "가용재고", "예약수량", "안전재고"}
    if header_set & inventory_keywords or "재고" in lower_text:
        return "inventory"

    cost_keywords = [
        "budget", "price", "cost", "amount", "quotation", "quote",
        "견적", "단가", "금액", "합계",
    ]
    if any(keyword in lower_text for keyword in cost_keywords):
        return "cost"

    bom_keywords = ["bom", "part", "material", "품번", "부품", "자재", "규격"]
    if any(keyword in lower_text for keyword in bom_keywords):
        return "bom"

    return "general"


def _append_xlsx_block(
    results: list[dict],
    text: str,
    sheet_name: str,
    row_start: int,
    row_end: int,
) -> None:
    results.append({
        "text": text,
        "source": {
            "source_type": "xlsx",
            "page_number": None,
            "sheet_name": sheet_name,
            "row_start": row_start,
            "row_end": row_end,
        }
    })


def _flush_grouped_rows(
    results: list[dict],
    grouped_rows: list[tuple[int, str]],
    sheet_name: str,
) -> None:
    if not grouped_rows:
        return

    _append_xlsx_block(
        results=results,
        text="\n".join(text for _, text in grouped_rows),
        sheet_name=sheet_name,
        row_start=grouped_rows[0][0],
        row_end=grouped_rows[-1][0],
    )


def _grouped_text_length(grouped_rows: list[tuple[int, str]]) -> int:
    return sum(len(text) for _, text in grouped_rows)


def parse_xlsx(path: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    results = []
    filename = os.path.basename(path)
    group_size = max(1, settings.xlsx_general_row_group_size)
    group_max_chars = max(1, settings.xlsx_group_max_chars)

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        header_idx = _detect_header_row(rows)
        headers = [_cell_to_text(v) for v in rows[header_idx]]
        sheet_type = _detect_sheet_type(filename, ws.title, headers)
        grouped_rows = []

        for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if _is_empty_row(row):
                continue

            values = [_cell_to_text(v) for v in row]

            pairs = []

            for col_idx, value in enumerate(values):
                if not value:
                    continue

                header = headers[col_idx] if col_idx < len(headers) else ""

                if header:
                    pairs.append(f"{header}={value}")
                else:
                    pairs.append(value)

            if not pairs:
                continue

            raw_row_text = " | ".join(v for v in values if v)
            structured_text = ", ".join(pairs)

            text = (
                f"파일명={filename}, "
                f"시트={ws.title}, "
                f"행={row_idx}, "
                f"{structured_text}. "
                f"원본행={raw_row_text}"
            )

            if sheet_type in {"inventory", "bom", "cost"} or group_size == 1:
                _append_xlsx_block(
                    results=results,
                    text=text,
                    sheet_name=ws.title,
                    row_start=row_idx,
                    row_end=row_idx,
                )
                continue

            grouped_rows.append((row_idx, text))

            if (
                len(grouped_rows) >= group_size
                or _grouped_text_length(grouped_rows) >= group_max_chars
            ):
                _flush_grouped_rows(results, grouped_rows, ws.title)
                grouped_rows = []

        _flush_grouped_rows(results, grouped_rows, ws.title)

    return results
