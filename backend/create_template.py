from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

output_path = "./data/templates/quotation_template.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "견적서"

# 열 너비
widths = {
    "A": 18,
    "B": 16,
    "C": 12,
    "D": 14,
    "E": 16,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

# 제목
ws.merge_cells("A1:E1")
ws["A1"] = "데모 견적서"
ws["A1"].font = Font(size=18, bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# 기본 정보
info_rows = [
    ("A3", "고객사명", "B3"),
    ("A4", "프로젝트명", "B4"),
    ("A5", "견적일자", "B5"),
]
for label_cell, label, value_cell in info_rows:
    ws[label_cell] = label
    ws[label_cell].font = Font(bold=True)
    ws[label_cell].fill = PatternFill("solid", fgColor="D9EAF7")
    ws[value_cell] = ""

# 품목 헤더
headers = ["품목명", "품번", "수량", "단가", "금액"]
for idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=8, column=idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="00897B")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 총액
ws["A20"] = "총액"
ws["A20"].font = Font(bold=True)
ws["E20"] = 0
ws["E20"].number_format = '#,##0"원"'

# 스타일
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=5):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

for row in range(1, 21):
    ws.row_dimensions[row].height = 24

wb.save(output_path)
print(f"template created: {output_path}")